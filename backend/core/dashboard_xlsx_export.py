"""Exportación Excel (.xlsx): tablas agregadas + gráfico raster (matplotlib).

Los datos coinciden con el dashboard/DW; la figura se genera en servidor (no es Chart.js).
"""
from __future__ import annotations

import io
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage


def _new_png_buffer() -> io.BytesIO:
    return io.BytesIO()


def _pie_fig(labels: list[str], values: list[float], title: str) -> io.BytesIO:
    fig, ax = plt.subplots(figsize=(5.2, 4.2))
    colors = ["#69d2ff", "#a78bfa", "#58f2b3", "#ffd166", "#ff6b9d", "#94a3b8"]
    ax.pie(
        values,
        labels=labels,
        autopct=lambda p: f"{p:.1f}%" if p > 6 else "",
        colors=[colors[i % len(colors)] for i in range(len(labels))],
        textprops={"fontsize": 9},
    )
    ax.set_title(title, fontsize=11, pad=10)
    buf = _new_png_buffer()
    fig.tight_layout()
    fig.savefig(buf, format="png", dpi=125, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


def _barh_fig(labels: list[str], values: list[float], title: str) -> io.BytesIO:
    n = len(labels)
    fig_h = max(3.0, min(10.5, 0.36 * n + 1.2))
    fig, ax = plt.subplots(figsize=(6.4, fig_h))
    y = list(range(n))
    ax.barh(y, values, color="#69d2ff", edgecolor="#4dbbe8", linewidth=0.4)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=8)
    ax.invert_yaxis()
    ax.set_title(title, fontsize=11)
    ax.set_xlabel("Total")
    buf = _new_png_buffer()
    fig.tight_layout()
    fig.savefig(buf, format="png", dpi=125, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


def _barv_fig(labels: list[str], values: list[float], title: str) -> io.BytesIO:
    n = len(labels)
    fig_w = max(6.5, min(15.0, 0.32 * n + 2))
    fig, ax = plt.subplots(figsize=(fig_w, 4.3))
    x = list(range(n))
    ax.bar(x, values, color="#69d2ff", edgecolor="#4dbbe8", linewidth=0.4)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=48, ha="right", fontsize=7)
    ax.set_title(title, fontsize=11)
    ax.set_ylabel("Total")
    buf = _new_png_buffer()
    fig.tight_layout()
    fig.savefig(buf, format="png", dpi=125, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


def _short_label(s: str, max_len: int = 34) -> str:
    if s is None:
        return "—"
    t = str(s)
    return t if len(t) <= max_len else f"{t[: max_len - 1]}…"


def panel_figure_png_buffer(panel: str, rows: list[dict[str, Any]]) -> io.BytesIO | None:
    if not rows:
        return None
    if panel == "canal":
        labels = [_short_label(str(r.get("canal__canal") or "—")) for r in rows]
        vals = [float(r.get("total") or 0) for r in rows]
        return _pie_fig(labels, vals, "Por canal")
    if panel == "clientes":
        labels = []
        for r in rows:
            nm = f"{r.get('cliente__nombre') or ''} {r.get('cliente__apellido') or ''}".strip()
            labels.append(_short_label(nm or str(r.get("cliente__id_cliente") or "")))
        vals = [float(r.get("total") or 0) for r in rows]
        return _barh_fig(labels, vals, "Top clientes")
    if panel == "dia":
        labels = []
        for r in rows:
            fc = r.get("fecha__fecha_completa")
            labels.append(fc.isoformat() if hasattr(fc, "isoformat") else str(fc or ""))
        vals = [float(r.get("total") or 0) for r in rows]
        return _barv_fig(labels, vals, "Por día civil")
    if panel == "dow":
        labels = [_short_label(str(r.get("fecha__nombre_dia") or "")) for r in rows]
        vals = [float(r.get("total") or 0) for r in rows]
        return _barv_fig(labels, vals, "Por día de la semana")
    if panel == "productos":
        labels = [_short_label(str(r.get("producto__nombre_producto") or "")) for r in rows]
        vals = [float(r.get("total") or 0) for r in rows]
        return _barh_fig(labels, vals, "Top productos")
    return None


def _write_panel_block(ws, panel: str, rows: list[dict[str, Any]], filter_summary: str) -> int:
    """Escribe tabla; devuelve fila donde colocar la imagen (o 0 si no hay figura)."""
    titles = {
        "canal": "Por canal",
        "clientes": "Top clientes",
        "dia": "Por día civil",
        "dow": "Por día de la semana",
        "productos": "Top productos",
    }
    title = titles.get(panel, panel)
    ws.append([title])
    ws.append(["Filtros aplicados:", filter_summary or "—"])
    ws.append([])
    if not rows:
        ws.append(["Sin datos para los filtros actuales."])
        return

    if panel == "canal":
        ws.append(["Canal", "Total"])
        for r in rows:
            ws.append([r.get("canal__canal"), r.get("total")])
        img_row_start = ws.max_row + 2
    elif panel == "clientes":
        ws.append(["Cliente id", "Nombre", "Apellido", "Segmento", "Total"])
        for r in rows:
            ws.append(
                [
                    r.get("cliente__id_cliente"),
                    r.get("cliente__nombre"),
                    r.get("cliente__apellido"),
                    r.get("cliente__segmento"),
                    r.get("total"),
                ]
            )
        img_row_start = ws.max_row + 2
    elif panel == "dia":
        ws.append(["Fecha", "Día", "Transacciones", "Total"])
        for r in rows:
            fc = r.get("fecha__fecha_completa")
            fds = fc.isoformat() if hasattr(fc, "isoformat") else str(fc or "")
            ws.append([fds, r.get("fecha__nombre_dia"), r.get("tx"), r.get("total")])
        img_row_start = ws.max_row + 2
    elif panel == "dow":
        ws.append(["Día", "Número ISO (día)", "Total"])
        for r in rows:
            ws.append([r.get("fecha__nombre_dia"), r.get("fecha__dia_semana"), r.get("total")])
        img_row_start = ws.max_row + 2
    elif panel == "productos":
        ws.append(["Producto id", "Nombre", "Total"])
        for r in rows:
            ws.append([r.get("producto__id_producto"), r.get("producto__nombre_producto"), r.get("total")])
        img_row_start = ws.max_row + 2
    else:
        ws.append([f"Panel no soportado: {panel}"])
        return

    png_buf = panel_figure_png_buffer(panel, rows)
    if png_buf:
        xl_img = XLImage(png_buf)
        scale = min(1.0, 560 / max(1, xl_img.width))
        xl_img.width = int(xl_img.width * scale)
        xl_img.height = int(xl_img.height * scale)
        ws.add_image(xl_img, f"A{img_row_start}")
    return


def build_dashboard_xlsx_bytes(
    *,
    panels: list[str],
    rows_map: dict[str, list],
    filter_summary: str,
) -> bytes:
    wb = Workbook()
    titles_sheet = {
        "canal": "Por canal",
        "clientes": "Top clientes",
        "dia": "Por día civil",
        "dow": "Por día semana",
        "productos": "Top productos",
    }

    for i, p in enumerate(panels):
        safe_title = str(titles_sheet.get(p, p))[:31]
        if i == 0:
            ws = wb.active
            if ws is None:
                raise RuntimeError("openpyxl: libro sin hoja activa")
            ws.title = safe_title
        else:
            ws = wb.create_sheet(title=safe_title)
        _write_panel_block(ws, p, list(rows_map.get(p, [])), filter_summary)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
